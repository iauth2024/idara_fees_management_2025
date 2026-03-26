import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Sum, F

def generate_excel(request):
    students = Student.objects.all()
    course = request.GET.get('course', '')
    branch = request.GET.get('branch', '')
    section = request.GET.get('section', '')
    months_paid = request.GET.get('months_paid', '')

    if months_paid.isdigit():
        months_paid = int(months_paid)
        students = students.annotate(total_paid=Sum('payment__amount')).filter(total_paid__gte=F('monthly_fees') * months_paid)

    wb = Workbook()
    ws = wb.active
    headers = ['Admission Number', 'Name', 'Course', 'Branch', 'Section', 'Monthly Fees', 'Total Fees', 'Total Paid', 'Total Due', 'Months Paid']
    ws.append(headers)

    for student in students:
        total_paid = student.payment_set.aggregate(total_paid=Sum('amount'))['total_paid'] or 0
        total_due = student.total_fees - total_paid
        months_paid_count = total_paid / student.monthly_fees if student.monthly_fees != 0 else 0

        row_data = [
            student.admission_number,
            student.name,
            student.course,
            student.branch,
            student.section,
            student.monthly_fees,
            student.total_fees,
            total_paid,
            total_due,
            months_paid_count
        ]
        ws.append(row_data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="students.xlsx"'
    wb.save(response)
    return response
##############################################################################################################

from django.contrib.auth import views as auth_views
from django.urls import reverse_lazy

class CustomPasswordResetView(auth_views.PasswordResetView):
    template_name = 'custom_password_reset.html'
    email_template_name = 'custom_password_reset_email.html'
    success_url = reverse_lazy('password_reset_done')

class CustomPasswordResetDoneView(auth_views.PasswordResetDoneView):
    template_name = 'custom_password_reset_done.html'

class CustomPasswordResetConfirmView(auth_views.PasswordResetConfirmView):
    template_name = 'custom_password_reset_confirm.html'
    success_url = reverse_lazy('password_reset_complete')

class CustomPasswordResetCompleteView(auth_views.PasswordResetCompleteView):
    template_name = 'custom_password_reset_complete.html'

##############################################################################################################


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Sum, Count, F, Case, When, IntegerField

from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import (
    F, Sum, Count, Case, When, IntegerField, Value, ExpressionWrapper, DecimalField
)
from django.db.models.functions import Coalesce
from django.shortcuts import render
import logging
from decimal import Decimal

from .models import Payment, Student

logger = logging.getLogger(__name__)

from django.db.models import Subquery, OuterRef, DecimalField, Value, Sum, Count, F, ExpressionWrapper, Q
from django.db.models.functions import Coalesce

@login_required
@staff_member_required
def summary(request):
    selected_organization = request.GET.get('organization', '')
    selected_year = request.GET.get('year', '')

    # Get distinct organizations and years from payments
    organizations = Payment.objects.values_list('organization', flat=True).distinct()
    years = Payment.objects.values_list('year', flat=True).distinct()

    # Base queryset for payments
    payments = Payment.objects.select_related('student', 'created_by').filter(student__isnull=False)
    
    # Apply filters to payments
    if selected_organization:
        payments = payments.filter(organization=selected_organization)
    if selected_year:
        payments = payments.filter(year=selected_year)

    # Get unique student IDs from filtered payments
    student_ids = payments.values_list('student_id', flat=True).distinct()
    
    # Annotate each student with their total fee and paid amount
    students = Student.objects.filter(id__in=student_ids).annotate(
        total_fee=F('monthly_fees') * 12,
        total_paid=Coalesce(
            Sum('payment__amount', filter=Q(payment__in=payments)),
            Value(0.0, output_field=DecimalField())
        )
    ).annotate(
        fee_due=F('total_fee') - F('total_paid')
    )

    # Overall totals
    total_students = students.count()
    total_fees = students.aggregate(total=Sum('total_fee'))['total'] or Decimal('0.0')
    collected_fees = students.aggregate(total=Sum('total_paid'))['total'] or Decimal('0.0')
    due_fees = total_fees - collected_fees
    fee_cleared_students = students.filter(fee_due__lte=0).count()

    # ===== BRANCH WISE TOTALS - CORRECTED =====
    # Get payment totals by branch directly from payments
    branch_payment_totals = (
        payments.values('student__branch')
        .annotate(
            total_paid=Coalesce(
                Sum('amount'),
                Value(Decimal('0.0'), output_field=DecimalField())
            ),
            number_of_students=Count('student', distinct=True),
        )
    )

    # Calculate due amounts by getting total_fee for each branch from students
    branch_list = []
    for item in branch_payment_totals:
        branch_name = item['student__branch'] or "N/A"
        
        # Get total fees for students in this branch
        students_in_branch = students.filter(branch=item['student__branch'])
        total_fee_for_branch = students_in_branch.aggregate(
            total=Sum('total_fee')
        )['total'] or Decimal('0.0')

        total_paid = item['total_paid']
        due_amount = total_fee_for_branch - total_paid

        branch_list.append({
            'student__branch': branch_name,
            'total': total_paid,
            'number_of_students': item['number_of_students'],
            'due_amount': due_amount
        })
    
    branch_wise_totals = sorted(branch_list, key=lambda x: x['student__branch'])

    # ===== COURSE WISE TOTALS - CORRECTED =====
    # Get payment totals by course directly from payments
    course_payment_totals = (
        payments.values('student__course')
        .annotate(
            total_paid=Coalesce(
                Sum('amount'),
                Value(Decimal('0.0'), output_field=DecimalField())
            ),
            number_of_students=Count('student', distinct=True),
        )
    )

    # Calculate due amounts by getting total_fee for each course from students
    course_list = []
    for item in course_payment_totals:
        course_name = item['student__course'] or "N/A"
        
        # Get total fees for students in this course
        students_in_course = students.filter(course=item['student__course'])
        total_fee_for_course = students_in_course.aggregate(
            total=Sum('total_fee')
        )['total'] or Decimal('0.0')

        total_paid = item['total_paid']
        due_amount = total_fee_for_course - total_paid

        course_list.append({
            'student__course': course_name,
            'total': total_paid,
            'number_of_students': item['number_of_students'],
            'due_amount': due_amount
        })
    
    course_wise_totals = sorted(course_list, key=lambda x: x['student__course'])

    # ===== SECTION WISE TOTALS - CORRECTED =====
    # Get payment totals by section directly from payments
    section_payment_totals = (
        payments.values('student__section')
        .annotate(
            total_paid=Coalesce(
                Sum('amount'),
                Value(Decimal('0.0'), output_field=DecimalField())
            ),
            number_of_students=Count('student', distinct=True),
        )
    )

    # Calculate due amounts by getting total_fee for each section from students
    section_list = []
    for item in section_payment_totals:
        section_name = item['student__section'] or "N/A"
        
        # Get total fees for students in this section
        students_in_section = students.filter(section=item['student__section'])
        total_fee_for_section = students_in_section.aggregate(
            total=Sum('total_fee')
        )['total'] or Decimal('0.0')

        total_paid = item['total_paid']
        due_amount = total_fee_for_section - total_paid

        section_list.append({
            'student__section': section_name,
            'total': total_paid,
            'number_of_students': item['number_of_students'],
            'due_amount': due_amount
        })
    
    section_wise_totals = sorted(section_list, key=lambda x: x['student__section'])

    # ===== USER WISE TOTALS - This one is correct as is =====
    user_wise_totals = (
        payments.values('created_by__username')
        .annotate(
            total=Coalesce(
                Sum('amount', output_field=DecimalField(max_digits=10, decimal_places=2)),
                Value(Decimal('0.0'), output_field=DecimalField(max_digits=10, decimal_places=2))
            ),
            number_of_entries=Count('id')
        )
        .order_by('created_by__username')
    )

    # Convert to list of dicts with proper field names
    user_list = []
    for item in user_wise_totals:
        user_list.append({
            'created_by__username': item['created_by__username'] or "N/A",
            'total': item['total'],
            'number_of_entries': item['number_of_entries']
        })
    user_wise_totals = user_list

    context = {
        'total_students': total_students,
        'total_fees': total_fees,
        'collected_fees': collected_fees,
        'due_fees': due_fees,
        'fee_cleared_students': fee_cleared_students,
        'branch_wise_totals': branch_wise_totals,
        'user_wise_totals': user_wise_totals,
        'class_wise_totals': section_wise_totals,
        'course_wise_totals': course_wise_totals,
        'organizations': organizations,
        'years': years,
        'selected_organization': selected_organization,
        'selected_year': selected_year,
    }

    return render(request, 'summary.html', context)

##############################################################################################################


from django.shortcuts import render
from django.contrib.auth.decorators import login_required

@login_required
def user_payments(request):
    user_payments = Payment.objects.filter(created_by=request.user).order_by('-date')
    context = {
        'user_payments': user_payments,
    }
    return render(request, 'user_payments.html', context)

##############################################################################################################

from django.shortcuts import render
from django.db.models import Sum

def student_payment_report(request):
    if request.method == 'POST':
        admission_number = request.POST.get('admission_number')
        try:
            student = Student.objects.get(admission_number=admission_number)
            payments = Payment.objects.filter(student=student).order_by('-date')

            monthly_fee = student.monthly_fees
            total_fee = student.total_fees
            total_paid = Payment.objects.filter(student=student).aggregate(Sum('amount'))['amount__sum'] or 0
            fee_due = total_fee - total_paid

            return render(request, 'student_payment_report.html', {
                'student': student,
                'payments': payments,
                'monthly_fee': monthly_fee,
                'total_fee': total_fee,
                'total_paid': total_paid,
                'fee_due': fee_due
            })
        except Student.DoesNotExist:
            return render(request, 'student_payment_report.html', {'error_message': 'Student not found'})
    else:
        return render(request, 'student_payment_report.html')

##############################################################################################################

import pandas as pd
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from .forms import UploadFileForm
from .models import Student
from decimal import Decimal, InvalidOperation


from decimal import Decimal, InvalidOperation
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
import pandas as pd

import pandas as pd
from decimal import Decimal, InvalidOperation
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import UploadFileForm
from .models import Student

@login_required
@user_passes_test(lambda u: u.is_staff)
def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            if file.name.endswith('.xlsx'):  # Check for Excel file type
                df = pd.read_excel(file)
                success_students = []
                failed_students = []
                already_exists_students = []

                # Ensure expected columns are present
                required_columns = [
                    'Admission Number', 'Name', 'Father Name', 'Phone', 'Course', 'Section', 'Branch', 'Monthly Donation', 'Student Type'
                ]
                if not all(col in df.columns for col in required_columns):
                    return render(request, 'upload_result.html', {
                        'error': 'The uploaded file is missing one or more required columns.'
                    })

                for index, row in df.iterrows():
                    admission_number = row['Admission Number']
                    # Ensure phone is treated as a string and strip any whitespace
                    phone = str(row['Phone']).strip()

                    # Log or print the phone number to debug
                    print(f"Processing phone number: {phone}")

                    # Ensure monthly donations are handled correctly as Decimal
                    try:
                        monthly_donation = Decimal(row['Monthly Donation']).quantize(Decimal('1.'))
                    except (ValueError, InvalidOperation):
                        failed_students.append({
                            'admission_number': admission_number,
                            'name': row['Name'],
                            'reason': 'Invalid Monthly Donation'
                        })
                        continue  # Skip to the next row if there's an error

                    # Check if the student already exists by admission number
                    if Student.objects.filter(admission_number=admission_number).exists():
                        already_exists_students.append({
                            'admission_number': admission_number,
                            'name': row['Name'],
                            'reason': 'Already Exists'
                        })
                    else:
                        # Check for duplicates based on other fields
                        if Student.objects.exclude(admission_number=admission_number).filter(
                                name=row['Name'],
                                phone=phone,  # Use the processed phone number
                                course=row['Course'],
                                section=row['Section'],
                                branch=row['Branch'],
                                monthly_fees=monthly_donation,
                                student_type=row['Student Type']
                        ).exists():
                            failed_students.append({
                                'admission_number': admission_number,
                                'name': row['Name'],
                                'reason': 'Duplicate Number'
                            })
                        else:
                            # Create a new student record and save
                            student = Student(
                                admission_number=admission_number,
                                name=row['Name'],
                                father_name=row['Father Name'],  # Add Father Name
                                phone=phone,  # Use the processed phone number
                                course=row['Course'],
                                section=row['Section'],
                                branch=row['Branch'],
                                monthly_fees=monthly_donation,
                                student_type=row['Student Type']
                            )
                            student.save()
                            success_students.append(student)

                # Calculate totals
                total_students_to_upload = len(df)
                already_exists_no = len(already_exists_students)
                failed_no = len(failed_students)
                success_no = total_students_to_upload - failed_no - already_exists_no
                newly_added_no = success_no

                return render(request, 'upload_result.html', {
                    'total_students_to_upload': total_students_to_upload,
                    'already_exists_no': already_exists_no,
                    'failed_no': failed_no,
                    'success_no': success_no,
                    'newly_added_no': newly_added_no,
                    'already_exists_students': already_exists_students,
                    'failed_students': failed_students,
                    'success_students': success_students,
                })
    else:
        form = UploadFileForm()
    return render(request, 'upload_file.html', {'form': form})





#############################################################################################################





import pandas as pd
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import PaymentUploadForm  # Ensure you have this import for your form
from .models import Payment, Student  # Ensure you have the correct model imports


import pandas as pd
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import PaymentUploadForm  # Ensure you have this import for your form
from .models import Payment, Student  # Ensure you have the correct model imports

@login_required
@user_passes_test(lambda u: u.is_staff)  # Ensure only admin users can access this view
def upload_payments(request):
    if request.method == 'POST':
        form = PaymentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)
            
            for index, row in df.iterrows():
                receipt_no = row['Receipt No']
                student_admission_no = row['Student Admission No']
                amount = row['Amount']
                date = row['Date']
                created_by_username = row['Created By']
                
                try:
                    student = Student.objects.get(admission_number=student_admission_no)
                except Student.DoesNotExist:
                    print(f"Student with admission number {student_admission_no} does not exist.")
                    continue
                
                try:
                    created_by = User.objects.get(username=created_by_username)
                except User.DoesNotExist:
                    print(f"User '{created_by_username}' does not exist.")
                    continue

                # Check for duplicate receipt_no
                if Payment.objects.filter(receipt_no=receipt_no).exists():
                    print(f"Receipt No {receipt_no} already exists. Skipping...")
                    continue  # Skip this row

                # Create and save Payment instance
                payment = Payment(
                    receipt_no=receipt_no,
                    student=student,
                    amount=amount,
                    date=date,
                    created_by=created_by
                )
                payment.save()
            
            return HttpResponse('Payments uploaded successfully')
    else:
        form = PaymentUploadForm()
    return render(request, 'upload_payments.html', {'form': form})


#############################################################################################################


from django.db.models import Sum
from .models import Student

def calculate_total_due(student):
    # Placeholder function to calculate total due for a student
    total_fees = student.total_fees  # Assuming you have a field named 'total_fees' in your Student model
    total_paid = student.payment_set.aggregate(total_paid=Sum('amount'))['total_paid'] or 0
    total_due = total_fees - total_paid
    return total_due
from django.http import JsonResponse
from .models import Student
from django.db.models import Sum

def get_student_details(request):
    if request.method == 'GET':
        admission_number = request.GET.get('admission_number')
        try:
            student = Student.objects.get(admission_number=admission_number)
            total_fees_paid = sum(payment.amount for payment in student.payment_set.all())
            total_due = student.total_fees - total_fees_paid
            months_paid = total_fees_paid / student.monthly_fees
            # Round off to 2 decimal places for monetary values
            total_fees_paid = round(total_fees_paid)
            total_due = round(total_due)
            # Round off to 1 decimal place for months_paid
            months_paid = round(months_paid, 1)
            data = {
                'name': student.name,
                'phone': student.phone,
                'course': student.course,
                'branch': student.branch,
                'monthly_fees': int(student.monthly_fees),  # Convert to integer to remove decimal places
                'total_fees': int(student.total_fees),  # Convert to integer to remove decimal places
                'total_paid': total_fees_paid,
                'total_due': total_due,
                'months_paid': months_paid
            }
            return JsonResponse({'success': True, 'student': data})
        except Student.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Admission number not found'}, status=404)



################################################################################################################






from django.shortcuts import render
from .models import Student

def download_students_page(request):
    # Get distinct values for dropdown filters
    courses = Student.objects.values_list('course', flat=True).distinct()
    branches = Student.objects.values_list('branch', flat=True).distinct()
    
    return render(request, 'download_students.html', {
        'courses': courses,
        'branches': branches,
    })
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from .models import Student

from io import BytesIO
from openpyxl import Workbook
from django.http import HttpResponse
from .models import Student

# views.py
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from .models import Student

def download_students(request):
    # Fetch all students from the database
    students = Student.objects.all()

    # Create a workbook and select the active worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Students'

    # Write the header row (including Father Name)
    headers = [
        'Admission Number', 'Name', 'Father Name', 'Phone', 'Course', 'Section', 'Branch', 'Monthly Donation', 'Student Type'
    ]
    sheet.append(headers)

    # Write the student data rows (including Father Name)
    for student in students:
        sheet.append([
            student.admission_number,  # Admission Number
            student.name,              # Name
            student.father_name,       # Father Name
            student.phone,             # Phone
            student.course,            # Course
            student.section,           # Section
            student.branch,            # Branch
            student.monthly_fees,      # Monthly Donation (variable name remains same)
            student.student_type       # Student Type
        ])

    # Create the HTTP response with Excel file content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="students.xlsx"'

    # Save the workbook to the response
    with BytesIO() as buffer:
        workbook.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())  # Write the content of the workbook to the response

    return response

from django.shortcuts import render, redirect
from django.urls import reverse

def receipt_number_input(request):
    if request.method == 'POST':
        receipt_no = request.POST.get('receipt_no')
        if receipt_no:
            return redirect(reverse('print_receipt', kwargs={'receipt_no': receipt_no}))
    return render(request, 'receipt_number_input.html')
from django.shortcuts import render, get_object_or_404
from .models import Payment

def print_receipt(request, receipt_no):
    # Retrieve the receipt using the provided receipt number
    receipt = get_object_or_404(Payment, receipt_no=receipt_no)
    
    # Assuming there's a relation to the Student model
    student = getattr(receipt, 'student', None)  # Adjust this according to your actual model relationships

    # Pass the receipt and student information to the template
    context = {
        'receipt': receipt,
        'student': student,
    }
    
    return render(request, 'print_receipt.html', context)



#############################################################################################################

from django.shortcuts import render
from django.db.models import Sum
from .models import Payment

def student_receipt_list(request):
    selected_organization = request.GET.get('organization', '')
    selected_year = request.GET.get('year', '')

    # Get distinct values for organizations and years
    organizations = Payment.objects.values_list('organization', flat=True).distinct()
    years = Payment.objects.values_list('year', flat=True).distinct()
    years = sorted(set(years), reverse=True)  # Ensure years are sorted and unique

    # Define the allowed receipt types and their thresholds
    allowed_receipt_types = {
        'Library Fee': 3600,
        'Hostel Fee': 12000,
        'Academic Fee': 12000,
        'Mess Fee': 48000,
        'Admission Fee': 2000,
        'Examination Fee': 400,
        'Stationary Fee': 12000,
    }

    # Filter payments based on selected organization and year
    payments = Payment.objects.all()
    if selected_organization:
        payments = payments.filter(organization=selected_organization)
    if selected_year:
        payments = payments.filter(year=selected_year)

    # Filter receipt types based on allowed receipt types
    receipt_types = [rt for rt in payments.values_list('receipt_type', flat=True).distinct() if rt in allowed_receipt_types]

    # Organize data by name and receipt type
    student_data = {}
    for payment in payments:
        student_name = payment.name
        if student_name.lower() == "atiya":
            continue  # Skip "atiya"

        # Only include data for specified receipt types
        if payment.receipt_type in allowed_receipt_types:
            if student_name not in student_data:
                student_data[student_name] = {receipt_type: 0 for receipt_type in receipt_types}
                student_data[student_name]['name'] = student_name
            student_data[student_name][payment.receipt_type] += payment.amount

    # Filter out names that don't meet the fee criteria
    filtered_data = {}
    for name, amounts in student_data.items():
        meets_criteria = any(amounts.get(rt, 0) > threshold for rt, threshold in allowed_receipt_types.items())
        if meets_criteria:
            filtered_data[name] = amounts

    context = {
        'student_data': filtered_data,
        'receipt_types': receipt_types,
        'organizations': organizations,
        'years': years,
        'selected_organization': selected_organization,
        'selected_year': selected_year,
        'allowed_receipt_types': allowed_receipt_types,  # Pass allowed receipt types and thresholds to the template
    }

    return render(request, 'student_receipt_list.html', context)

#################################################################################################

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm

from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect

def login_view(request):

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            # Leave group
            if user.groups.filter(name="leave").exists():
                return redirect('/leave')

            # Fee Collector group
            elif user.groups.filter(name="Fee Collector").exists():
                return redirect('/homepage/')

            # Default
            return redirect('/homepage/')

        else:
            return render(request, "login.html", {
                "error_message": "Invalid username or password"
            })

    return render(request, "login.html")

from django.shortcuts import redirect
from django.contrib.auth import logout

def logout_view(request):
    if request.method == 'POST':
        logout(request)
        return redirect('login')  # Redirect to login page after logout
    else:
        return redirect('login')  # Handle GET requests by redirecting to the login page


#################################################################################################

from django.shortcuts import render
from django.db.models import Sum
from .models import Payment

@login_required
def homepage(request):
    # Calculate the total collected amount by the logged-in user
    total_collected = Payment.objects.filter(created_by=request.user).aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'total_collected': total_collected,
    }
    return render(request, 'homepage.html', context)

################################################################################################################


import csv
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from .models import Payment

import openpyxl
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Payment

@login_required
@staff_member_required  # Ensures only staff members can access the view
def download_payments(request):
    payments = Payment.objects.all()

    # Create an Excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Payments"

    # Write header row
    headers = ['Receipt No', 'Student Admission No', 'Student Name', 'Amount', 'Date', 'Created By']
    sheet.append(headers)

    # Write data rows
    for payment in payments:
        sheet.append([
            payment.receipt_no,
            payment.student.admission_number if payment.student else '',
            payment.student.name if payment.student else '',
            payment.amount,
            payment.date.strftime('%Y-%m-%d') if payment.date else '',  # Format date to 'YYYY-MM-DD'
            payment.created_by.username if payment.created_by else ''  # Use username if available, else empty string
        ])

    # Prepare the response with appropriate content type for Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payments.xlsx"'

    # Save the workbook to the response
    workbook.save(response)

    return response

#########################################################################################################################
from django.shortcuts import render
from django.db.models import Q


from django.shortcuts import render
from django.db.models import Q
from fees_collection.models import Payment  # Import the correct model

def payment_list(request):
    query = request.GET.get('q', '')  # Get the search query from the request
    payments = Payment.objects.all().order_by('-date')  # Correctly refer to the model Payment

    if query:  # If there is a search query
        # Filter payments based on receipt no, student name, or payment method
        payments = payments.filter(
            Q(receipt_no__icontains=query) |  # Correct field name: receipt_no
            Q(student__name__icontains=query) |   # Filter by student name
            Q(payment_method__icontains=query)      # Filter by payment method (optional)
        )

    return render(request, 'payment_list.html', {'payments': payments, 'query': query})





#############################################################################################################################

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import PaymentForm
from .models import Payment, Student
from django.contrib.auth.models import User
from datetime import date
from django.db import IntegrityError


  # Import your models
# Add other necessary imports like User if needed

@login_required
def make_payment(request):
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            admission_number = form.cleaned_data['admission_number']
            try:
                student = Student.objects.get(admission_number=admission_number)
            except Student.DoesNotExist:
                messages.error(request, "Student not found with the provided admission number.")
                return render(request, 'make_payment.html', {'form': form})

            total_due = calculate_total_due(student)
            amount_paid = form.cleaned_data['amount_paid']
            payment_method = form.cleaned_data['payment_method']  # New field for payment method
            payment_date = form.cleaned_data['date']  # Capture the date field

            if 0 < amount_paid <= total_due:
                receipt_number = form.cleaned_data['receipt_number']
                created_by = request.user
                
                try:
                    Payment.objects.create(
                        student=student,
                        amount=amount_paid,
                        receipt_no=receipt_number,
                        created_by=created_by,
                        date=payment_date,
                        payment_method=payment_method,
                    )
                    messages.success(request, "Payment successfully recorded!")
                    return redirect(reverse('payment_success') + f'?admission_number={admission_number}')
                except IntegrityError:
                    messages.error(request, "A payment with the same receipt number already exists. Please enter a unique receipt number.")
            else:
                messages.error(request, "You cannot make 0 payment or more than the total due amount. Please pay the correct amount.")
    else:
        form = PaymentForm()

    return render(request, 'make_payment.html', {'form': form})

#########################################################################################################################3

from django.shortcuts import render
from .models import Payment, Student

def payment_success(request):
    admission_number = request.GET.get('admission_number')
    try:
        student = Student.objects.get(admission_number=admission_number)
        payment_details = Payment.objects.filter(student=student).order_by('-date')
        return render(request, 'payment_success.html', {
            'student': student,
            'payment_details': payment_details
        })
    except Student.DoesNotExist:
        return render(request, 'payment_success.html', {
            'error_message': 'No payment found for the provided admission number'
        })




####################################################################################################################


from django.shortcuts import render
from django.db.models import Sum
from .models import Student, Payment
from decimal import Decimal

from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import render
from .models import Student, Payment

from django.db.models import Sum, DecimalField
from django.db.models.functions import Coalesce

@login_required
@staff_member_required
def reports(request):

    course = request.GET.get('course', '')
    branch = request.GET.get('branch', '')
    section = request.GET.get('section', '')
    months_due = request.GET.get('months_due', '')
    
    # -------- MONTH FILTERS --------
    # Get selected months from request (can be multiple)
    selected_months_raw = request.GET.getlist('months')  # e.g., ['4', '5', '6']
    
    # Month choices for dropdown
    MONTH_CHOICES = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    
    # Convert selected months to integers for calculations, but keep strings for template comparison
    selected_months_int = []
    for m in selected_months_raw:
        if m.isdigit():
            selected_months_int.append(int(m))
    
    # Calculate months passed based on selected months
    if selected_months_int:
        selected_months_int.sort()
        months_passed = len(selected_months_int)
    else:
        # Default to current month if no selection
        current_month = date.today().month
        selected_months_int = list(range(1, current_month + 1))
        months_passed = current_month
    
    # For template: keep as strings for checkbox comparison
    selected_months_str = [str(m) for m in selected_months_int]
    # ------------------------------------

    students = Student.objects.all()

    # Dropdown values
    all_branches = Student.objects.exclude(branch__isnull=True).exclude(branch='')\
        .values_list('branch', flat=True).distinct().order_by('branch')

    all_courses = Student.objects.exclude(course__isnull=True).exclude(course='')\
        .values_list('course', flat=True).distinct().order_by('course')

    all_sections = Student.objects.exclude(section__isnull=True).exclude(section='')\
        .values_list('section', flat=True).distinct().order_by('section')

    # Organized filter data
    organized_data = {}
    for branch_item in all_branches:
        branch_courses = {}
        branch_students = Student.objects.filter(branch=branch_item)

        for course_item in branch_students.exclude(course__isnull=True).exclude(course='')\
                .values_list('course', flat=True).distinct():

            sections = list(branch_students.filter(course=course_item)
                .exclude(section__isnull=True)
                .exclude(section='')
                .values_list('section', flat=True)
                .distinct())

            if sections:
                branch_courses[course_item] = sections

        if branch_courses:
            organized_data[branch_item] = branch_courses

    # Apply filters
    if branch:
        students = students.filter(branch=branch)

    if course:
        students = students.filter(course=course)

    if section:
        students = students.filter(section=section)

    # -------- PAYMENT ANNOTATIONS --------
    students = students.annotate(
        total_paid=Coalesce(
            Sum('payment__amount'),
            Decimal('0.0'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    ).annotate(

        expected_fee=ExpressionWrapper(
            F('monthly_fees') * months_passed,
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )

    ).annotate(

        total_due=ExpressionWrapper(
            F('expected_fee') - F('total_paid'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )

    ).annotate(

        payment_percentage=Case(
            When(expected_fee__gt=0,
                then=ExpressionWrapper(
                    (F('total_paid') * 100.0) / F('expected_fee'),
                    output_field=DecimalField(max_digits=5, decimal_places=2)
                )),
            default=Value(0.0),
            output_field=DecimalField(max_digits=5, decimal_places=2)
        )
    )

    # -------- MONTHS DUE FILTER --------
    if months_due and months_due.isdigit():

        months_due_int = int(months_due)

        students = students.annotate(

            months_paid=ExpressionWrapper(
                F('total_paid') / F('monthly_fees'),
                output_field=DecimalField(max_digits=5, decimal_places=2)
            )

        ).annotate(

            months_due_calc=ExpressionWrapper(
                Value(months_passed) - F('months_paid'),
                output_field=DecimalField(max_digits=5, decimal_places=2)
            )

        ).filter(months_due_calc__gte=months_due_int)

    # -------- STATUS COUNTS --------
    # MOVED THIS SECTION BEFORE SUMMARY CALCULATIONS
    fully_paid = students.filter(payment_percentage__gte=99.99).count()
    partial_paid = students.filter(payment_percentage__gt=0, payment_percentage__lt=99.99).count()
    no_payment = students.filter(payment_percentage=0).count()

    # -------- SUMMARY --------
    total_students = students.count()

    total_collected = students.aggregate(
        total=Coalesce(Sum('total_paid'), Decimal('0.0'))
    )['total']

    total_expected = students.aggregate(
        total=Coalesce(Sum('expected_fee'), Decimal('0.0'))
    )['total']

    total_due_amount = total_expected - total_collected

    # Calculate percentages
    overall_collection_percentage = (total_collected / total_expected * 100) if total_expected > 0 else 0
    total_due_percentage = (total_due_amount / total_expected * 100) if total_expected > 0 else 0
    
    fully_paid_percentage = (fully_paid / total_students * 100) if total_students > 0 else 0
    partial_paid_percentage = (partial_paid / total_students * 100) if total_students > 0 else 0
    no_payment_percentage = (no_payment / total_students * 100) if total_students > 0 else 0

    # -------- STUDENT DATA --------
    additional_info = []

    for student in students:

        monthly_fees = Decimal(str(student.monthly_fees or 0))
        total_paid = Decimal(str(student.total_paid or 0))
        expected_fee = Decimal(str(student.expected_fee or 0))
        total_due = expected_fee - total_paid

        months_paid_calc = (total_paid / monthly_fees) if monthly_fees > 0 else Decimal('0')

        if student.payment_percentage >= Decimal('99.99'):
            status = "Fully Paid"
            status_class = "success"
        elif student.payment_percentage == Decimal('0.0'):
            status = "No Payment"
            status_class = "danger"
        else:
            status = "Partial Payment"
            status_class = "warning"

        additional_info.append({
            'student': student,
            'admission_number': student.admission_number,
            'student_name': student.name,
            'father_name': student.father_name,
            'branch': student.branch,
            'course': student.course,
            'section': student.section,
            'monthly_fees': monthly_fees,
            'expected_fee': expected_fee,
            'total_paid': total_paid,
            'total_due': total_due,
            'months_paid': months_paid_calc,
            'payment_percentage': student.payment_percentage,
            'status': status,
            'status_class': status_class,
            'months_covered': selected_months_int,
        })

    additional_info.sort(key=lambda x: (-float(x['total_due']), x['student_name']))

    context = {
        'additional_info': additional_info,
        'organized_data': organized_data,
        'all_branches': all_branches,
        'all_courses': all_courses,
        'all_sections': all_sections,
        'selected_branch': branch,
        'selected_course': course,
        'selected_section': section,
        'selected_months_due': months_due,
        
        # Month filter data
        'month_choices': MONTH_CHOICES,
        'selected_months': selected_months_str,  # Send as strings for template
        'selected_months_int': selected_months_int,  # Send as ints for calculations if needed
        'months_passed': months_passed,

        'total_students': total_students,
        'total_collected': total_collected,
        'total_expected': total_expected,
        'total_due_amount': total_due_amount,
        
        # Percentages
        'overall_collection_percentage': overall_collection_percentage,
        'total_due_percentage': total_due_percentage,
        'fully_paid_percentage': fully_paid_percentage,
        'partial_paid_percentage': partial_paid_percentage,
        'no_payment_percentage': no_payment_percentage,

        'fully_paid': fully_paid,
        'partial_paid': partial_paid,
        'no_payment': no_payment,
    }

    return render(request, 'reports.html', context)
######################################################################################################################


from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa # type: ignore
from .models import Student, Payment
from django.db.models import Sum
from django.templatetags.static import static

def generate_pdf(request):
    # Fetch and filter students
    students = Student.objects.all()
    course = request.GET.get('course', '')
    branch = request.GET.get('branch', '')
    section = request.GET.get('section', '')
    months_paid = request.GET.get('months_paid', '')

    if course:
        students = students.filter(course=course)
    if branch:
        students = students.filter(branch=branch)
    if section:
        students = students.filter(section=section)
    if months_paid.isdigit():
        months_paid = int(months_paid)
        students = [student for student in students if (Payment.objects.filter(student=student).aggregate(Sum('amount'))['amount__sum'] or 0) / student.monthly_fees >= months_paid]

    # Prepare additional student information for the context
    additional_info = []
    for student in students:
        total_paid = Payment.objects.filter(student=student).aggregate(Sum('amount'))['amount__sum'] or 0
        total_due = student.total_fees - total_paid
        months_paid_count = total_paid / student.monthly_fees if student.monthly_fees != 0 else 0
        additional_info.append({
            'student': student,
            'monthly_fees': student.monthly_fees,
            'total_fees': student.total_fees,
            'total_paid': total_paid,
            'total_due': total_due,
            'months_paid': months_paid_count,
        })

    # Context for template rendering
    context = {
        'additional_info': additional_info,
    }

    # Load template and render it with context
    template_path = 'pdf_template.html'
    template = get_template(template_path)
    html = template.render(context)

    # Prepare response for PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="report.pdf"'

    # Add a callback for static files to ensure font access
    def link_callback(uri, rel):
        if uri.startswith("static/"):
            return static(uri)
        return uri

    # Generate PDF with Unicode support (UTF-8 encoding)
    pisa_status = pisa.CreatePDF(
        html, dest=response, encoding='UTF-8', link_callback=link_callback
    )

    # Error handling for PDF generation
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


######################################################################################################################

from django.shortcuts import render, redirect, get_object_or_404
from .models import Student
from .forms import StudentEditForm

from django.shortcuts import render
from .models import Student
from django.db.models import Q  # Import Q for complex queries

def student_list(request):
    query = request.GET.get('q', '')  # Get the search query from the request
    students = Student.objects.all()  # Start with all students

    if query:  # If there is a search query
        # Filter students based on admission number, phone, or name
        students = students.filter(
            Q(admission_number__icontains=query) | 
            Q(phone__icontains=query) | 
            Q(name__icontains=query)
        )

    return render(request, 'student_list.html', {'students': students, 'query': query})


from django.contrib import messages  # Import Django messages

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import Student
from .forms import StudentEditForm

def edit_student(request, admission_number):
    """View to edit student details using admission number."""
    student = get_object_or_404(Student, admission_number=admission_number)

    if request.method == 'POST':
        form = StudentEditForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student details updated successfully.')
            return redirect('student_list')
    else:
        form = StudentEditForm(instance=student)

    return render(request, 'edit_student.html', {'form': form})

####################################################################################################################

from decimal import Decimal
from datetime import date
import json
import traceback
import requests

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.db import models
from django.db.models import Sum, F, DecimalField, ExpressionWrapper, Value
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator

from fees_collection.models import Student, Payment, SMSCampaign
from .sms_utils import SMSAPI
from .sms_templates import SMSTemplates




@login_required
@staff_member_required
def fee_due_reminders(request):

    branch = request.GET.get('branch', '')
    course = request.GET.get('course', '')
    section = request.GET.get('section', '')
    months_due = request.GET.get('months_due', '')

    # ✅ MONTH LOGIC (LIKE REPORTS)
    if months_due and months_due.isdigit():
        months_passed = int(months_due)
    else:
        months_passed = date.today().month  # default

    students = Student.objects.all()

    # ✅ FILTERS
    if branch:
        students = students.filter(branch=branch)
    if course:
        students = students.filter(course=course)
    if section:
        students = students.filter(section=section)

    # ✅ CALCULATIONS (SAME AS REPORTS)
    students = students.annotate(
        total_paid=Coalesce(Sum('payment__amount'), Decimal('0'))
    ).annotate(
        expected_fee=ExpressionWrapper(
            F('monthly_fees') * months_passed,
            output_field=DecimalField()
        )
    ).annotate(
        total_due=ExpressionWrapper(
            F('expected_fee') - F('total_paid'),
            output_field=DecimalField()
        )
    )

    # ✅ ONLY SHOW DUE STUDENTS
    students = students.filter(total_due__gt=0)

    # ✅ DROPDOWNS (FOR FILTER UI)
    branches = Student.objects.values_list('branch', flat=True).distinct()
    courses = Student.objects.values_list('course', flat=True).distinct()
    sections = Student.objects.values_list('section', flat=True).distinct()

    context = {
        'students': students,
        'branches': branches,
        'courses': courses,
        'sections': sections,
        'selected_branch': branch,
        'selected_course': course,
        'selected_section': section,
        'months_due': months_passed,
    }

    return render(request, 'fee_due_reminders.html', context)


import json
import requests
from django.http import JsonResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required



@login_required
@staff_member_required
def send_fee_due_sms(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        data = json.loads(request.body)
        student_ids = data.get("student_ids", [])
        months_due = int(data.get("months_due", 3))

        if not student_ids:
            return JsonResponse({"error": "No students selected"}, status=400)

        students = Student.objects.filter(id__in=student_ids)

        results = []

        for student in students:
            if not student.phone:
                continue

            # 🔹 Calculate paid
            if hasattr(student, 'payments'):
                total_paid = student.payments.aggregate(
                    total=models.Sum('amount')
                )['total'] or Decimal('0')
            else:
                total_paid = student.payment_set.aggregate(
                    total=models.Sum('amount')
                )['total'] or Decimal('0')

            monthly_fee = student.monthly_fees or Decimal('0')
            expected_amount = monthly_fee * months_due
            due_amount = expected_amount - total_paid

            if due_amount <= 0:
                continue

            admission_no = student.admission_number or 'N/A'

            # ✅ EXACT TEMPLATE MATCH (DLT REQUIRED)
            message = f"Fee due of Rs {int(due_amount)} for {student.name} (Adm No:{admission_no}) pending. Ignore if paid. Idara Ashraful Uloom Trust"

            url = "https://smslogin.co/v3/api.php"

            params = {
                "username": "AUTHYD1",
                "apikey": "333e1783fe5aae5fd76d",
                "senderid": "AUTHYD",  # ✅ must match DLT
                "mobile": student.phone,
                "message": message,
                "templateid": "1707177381044708917"  # ✅ your template ID
            }

            response = requests.get(url, params=params)

            results.append({
                "name": student.name,
                "mobile": student.phone,
                "due": int(due_amount),
                "response": response.text
            })

        return JsonResponse({
            "status": "success",
            "sent_count": len(results),
            "results": results
        })

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)